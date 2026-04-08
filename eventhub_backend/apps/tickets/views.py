from __future__ import annotations

from decimal import Decimal
from datetime import date as dt_date, datetime, time as dt_time

from django.core.exceptions import ValidationError as DjangoValidationError
from django_ratelimit.decorators import ratelimit
from django.shortcuts import get_object_or_404
from rest_framework import generics, permissions, status
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from apps.accounts.permissions import IsOrganizer, IsOrganizerRole
from apps.events.models import Event
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from .models import PromoCode, TicketType, RegistrationCategory, RegistrationQuestion, School, Course
from .serializers import (
    PromoCodeSerializer,
    PromoCodeValidateSerializer,
    TicketTypeCreateSerializer,
    TicketTypeSerializer,
    RegistrationCategorySerializer,
    RegistrationCategoryWriteSerializer,
    SchoolSerializer,
    CourseSerializer,
)
import requests

from .importers import import_courses_file, import_schools_file


class EventTicketTypesView(generics.ListAPIView):
    serializer_class = TicketTypeSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        slug = self.kwargs["slug"]
        event = get_object_or_404(Event, slug=slug, status="published")
        return event.ticket_types.filter(is_active=True)


class TicketTypeCreateView(generics.CreateAPIView):
    serializer_class = TicketTypeCreateSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrganizerRole]

    def perform_create(self, serializer):
        event = get_object_or_404(Event, slug=self.kwargs["slug"], organizer=self.request.user)
        reg_category = serializer.validated_data.get("registration_category")
        if reg_category and reg_category.event_id != event.id:
            raise ValidationError({"registration_category": "Category does not belong to this event."})
        serializer.save(event=event)


class TicketTypeUpdateView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TicketTypeCreateSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrganizer]
    lookup_url_kwarg = "ticket_id"
    lookup_field = "id"

    def get_queryset(self):
        event = get_object_or_404(Event, slug=self.kwargs["slug"], organizer=self.request.user)
        return event.ticket_types.all()

    def perform_update(self, serializer):
        event = get_object_or_404(Event, slug=self.kwargs["slug"], organizer=self.request.user)
        reg_category = serializer.validated_data.get("registration_category")
        if reg_category and reg_category.event_id != event.id:
            raise ValidationError({"registration_category": "Category does not belong to this event."})
        serializer.save()


class EventRegistrationSetupView(APIView):
    """
    Organizer-only: configure registration categories + questions for an event.
    GET returns all categories (active + inactive) with questions.
    POST upserts categories + nested questions.
    """
    permission_classes = [permissions.IsAuthenticated, IsOrganizerRole]

    def get(self, request, slug):
        event = get_object_or_404(Event, slug=slug, organizer=request.user)
        categories = event.registration_categories.prefetch_related("questions").all().order_by("sort_order")
        data = RegistrationCategorySerializer(categories, many=True).data
        return Response({"categories": data}, status=status.HTTP_200_OK)

    def post(self, request, slug):
        event = get_object_or_404(Event, slug=slug, organizer=request.user)
        payload = request.data or {}
        categories_data = payload.get("categories", [])
        if not isinstance(categories_data, list):
            raise ValidationError({"categories": "Expected a list of categories."})

        # Map existing categories by id and by type for upsert
        existing = {str(c.id): c for c in event.registration_categories.all()}
        existing_by_type = {c.category: c for c in event.registration_categories.all()}

        saved = []
        with transaction.atomic():
            for cat in categories_data:
                serializer = RegistrationCategoryWriteSerializer(data=cat)
                serializer.is_valid(raise_exception=True)
                validated = serializer.validated_data

                category_type = validated.get("category")
                if not category_type:
                    raise ValidationError({"category": "Category type is required."})

                # Force label for student/alumni to default values
                if category_type in ["student", "alumni"]:
                    validated["label"] = "Student" if category_type == "student" else "Alumni"
                    if category_type == "student":
                        validated["require_student_email"] = True
                        validated["require_admission_number"] = True
                else:
                    # Guest label fallback
                    if not validated.get("label"):
                        validated["label"] = "Guest"

                category_id = cat.get("id")
                instance = existing.get(str(category_id)) if category_id else existing_by_type.get(category_type)
                if instance:
                    for field, value in validated.items():
                        if field == "questions":
                            continue
                        setattr(instance, field, value)
                    instance.event = event
                    instance.save()
                else:
                    instance = RegistrationCategory.objects.create(event=event, **{
                        k: v for k, v in validated.items() if k != "questions"
                    })

                # Upsert questions
                questions_data = validated.get("questions", [])
                existing_questions = {str(q.id): q for q in instance.questions.all()}
                incoming_ids = {str(q.get("id")) for q in questions_data if q.get("id")}
                for q in questions_data:
                    q_id = q.get("id")
                    q_instance = existing_questions.get(str(q_id)) if q_id else None
                    if q_instance:
                        for field, value in q.items():
                            if field == "id":
                                continue
                            setattr(q_instance, field, value)
                        q_instance.save()
                    else:
                        RegistrationQuestion.objects.create(
                            category=instance,
                            label=q.get("label", ""),
                            field_type=q.get("field_type", "text"),
                            is_required=bool(q.get("is_required", False)),
                            options=q.get("options") or [],
                            sort_order=int(q.get("sort_order") or 0),
                        )

                # Delete removed questions
                if questions_data is not None:
                    if incoming_ids:
                        for qid, qobj in existing_questions.items():
                            if qid not in incoming_ids:
                                qobj.delete()
                    else:
                        # If no questions sent, remove all existing
                        for qobj in existing_questions.values():
                            qobj.delete()

                saved.append(instance)

        data = RegistrationCategorySerializer(saved, many=True).data
        return Response({"categories": data}, status=status.HTTP_200_OK)


class EventRegistrationPublicView(generics.ListAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = RegistrationCategorySerializer

    def get_queryset(self):
        slug = self.kwargs["slug"]
        event = get_object_or_404(Event, slug=slug, status="published")
        return event.registration_categories.filter(is_active=True).prefetch_related("questions").order_by("sort_order")


class SchoolListView(generics.ListCreateAPIView):
    serializer_class = SchoolSerializer
    pagination_class = None
    queryset = School.objects.filter(is_active=True).order_by("sort_order", "name")

    def get_permissions(self):
        if self.request.method == "GET":
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated(), IsOrganizerRole()]


class CourseListView(generics.ListCreateAPIView):
    serializer_class = CourseSerializer
    pagination_class = None

    def get_queryset(self):
        qs = Course.objects.select_related("school").filter(is_active=True).order_by("sort_order", "name")
        school_id = self.request.query_params.get("school")
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs

    def get_permissions(self):
        if self.request.method == "GET":
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated(), IsOrganizerRole()]


class SchoolUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsOrganizerRole]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            raise ValidationError({"file": "Upload file is required."})
        try:
            summary = import_schools_file(file)
        except DjangoValidationError as exc:
            raise ValidationError({"file": exc.messages[0]})
        return Response(summary, status=status.HTTP_200_OK)


class CourseUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsOrganizerRole]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            raise ValidationError({"file": "Upload file is required."})
        try:
            summary = import_courses_file(file)
        except DjangoValidationError as exc:
            raise ValidationError({"file": exc.messages[0]})
        return Response(summary, status=status.HTTP_200_OK)


class LocationSearchView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        query = (request.query_params.get("q") or "").strip()
        if not query:
            return Response({"results": []}, status=status.HTTP_200_OK)

        token = getattr(settings, "MAPBOX_ACCESS_TOKEN", "") or getattr(settings, "LOCATIONIQ_TOKEN", "")
        if not token:
            return Response({"results": []}, status=status.HTTP_200_OK)

        # Default to Mapbox if token present
        try:
            url = "https://api.mapbox.com/geocoding/v5/mapbox.places/{q}.json".format(q=requests.utils.quote(query))
            params = {
                "access_token": token,
                "autocomplete": "true",
                "limit": 5,
                "types": "place,locality,neighborhood,address",
            }
            resp = requests.get(url, params=params, timeout=6)
            resp.raise_for_status()
            data = resp.json()
            results = []
            for feat in data.get("features", []):
                place = feat.get("place_name", "")
                center = feat.get("center") or []
                ctx = feat.get("context", []) or []
                city = ""
                country = ""
                for item in ctx:
                    if item.get("id", "").startswith("place"):
                        city = item.get("text", "")
                    if item.get("id", "").startswith("country"):
                        country = item.get("text", "")
                results.append({
                    "label": place,
                    "city": city,
                    "country": country,
                    "lat": center[1] if len(center) > 1 else None,
                    "lng": center[0] if len(center) > 0 else None,
                })
            return Response({"results": results}, status=status.HTTP_200_OK)
        except Exception:
            return Response({"results": []}, status=status.HTTP_200_OK)


class PromoCodeManageView(generics.ListCreateAPIView):
    serializer_class = PromoCodeSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrganizer]

    def get_queryset(self):
        event = get_object_or_404(Event, slug=self.kwargs["slug"], organizer=self.request.user)
        return event.promo_codes.all()

    def perform_create(self, serializer):
        event = get_object_or_404(Event, slug=self.kwargs["slug"], organizer=self.request.user)
        serializer.save(event=event)


class PromoCodeBulkUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsOrganizer]

    @staticmethod
    def _normalize_header(value) -> str:
        return str(value or "").strip().lower().replace(" ", "_")

    @staticmethod
    def _pick_value(row: dict, *keys):
        for key in keys:
            if key in row and row[key] not in (None, ""):
                return row[key]
        return None

    @staticmethod
    def _parse_bool(value, default=True) -> bool:
        if value is None or str(value).strip() == "":
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "t"}:
            return True
        if text in {"0", "false", "no", "n", "f"}:
            return False
        raise ValueError("Invalid active flag.")

    @staticmethod
    def _parse_decimal(value):
        if value is None or str(value).strip() == "":
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        raw = str(value).strip().replace(",", "")
        if raw.endswith("%"):
            raw = raw[:-1].strip()
        return Decimal(raw)

    @staticmethod
    def _parse_expiry(value):
        if value is None or str(value).strip() == "":
            return None
        if isinstance(value, datetime):
            expiry_dt = value
        elif isinstance(value, dt_date):
            expiry_dt = datetime.combine(value, dt_time.min)
        else:
            raw = str(value).strip()
            expiry_dt = parse_datetime(raw)
            if not expiry_dt:
                parsed_date = parse_date(raw)
                if parsed_date:
                    expiry_dt = datetime.combine(parsed_date, dt_time.min)
        if not expiry_dt:
            raise ValueError("Invalid expiry date.")
        if timezone.is_naive(expiry_dt):
            expiry_dt = timezone.make_aware(expiry_dt, timezone.get_current_timezone())
        return expiry_dt

    def post(self, request, slug):
        event = get_object_or_404(Event, slug=slug, organizer=request.user)
        file = request.FILES.get("file")
        if not file:
            raise ValidationError({"file": "Upload file is required."})

        rows = []
        filename = file.name.lower()

        if filename.endswith(".csv"):
            content = file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                normalized = {self._normalize_header(k): v for k, v in row.items() if k}
                if any(str(val).strip() for val in normalized.values() if val is not None):
                    rows.append(normalized)
        elif filename.endswith(".xlsx"):
            try:
                import openpyxl
            except Exception:
                raise ValidationError({"file": "Excel upload requires openpyxl."})
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [self._normalize_header(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2):
                data = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    data[header] = row[idx].value
                if any(str(val).strip() for val in data.values() if val is not None):
                    rows.append(data)
        else:
            raise ValidationError({"file": "Upload a CSV or XLSX file."})

        if not rows:
            return Response({"total": 0, "created": 0, "updated": 0, "skipped": 0, "errors": []})

        created = 0
        updated = 0
        skipped = 0
        errors = []

        for index, row in enumerate(rows, start=2):
            try:
                code_raw = self._pick_value(row, "code", "promo_code", "promo", "coupon", "coupon_code")
                if not code_raw:
                    raise ValueError("Missing code.")
                code = str(code_raw).strip().upper().replace(" ", "")
                if not code:
                    raise ValueError("Missing code.")

                discount_type_raw = self._pick_value(row, "discount_type", "type")
                discount_type = str(discount_type_raw or "percent").strip().lower()
                if discount_type in {"percent", "percentage", "pct", "%"}:
                    discount_type = "percent"
                elif discount_type in {"fixed", "amount", "flat", "value"}:
                    discount_type = "fixed"
                else:
                    raise ValueError("Invalid discount type.")

                discount_value_raw = self._pick_value(row, "discount_value", "discount", "value", "amount")
                discount_value = self._parse_decimal(discount_value_raw)
                if discount_value is None:
                    raise ValueError("Missing discount value.")

                usage_limit_raw = self._pick_value(row, "usage_limit", "limit", "usage")
                usage_limit = None
                if usage_limit_raw not in (None, ""):
                    usage_limit = int(float(str(usage_limit_raw).strip()))
                    if usage_limit <= 0:
                        usage_limit = None

                expiry_raw = self._pick_value(row, "expiry", "expires", "expiry_date", "expires_at", "expiry_at")
                expiry = self._parse_expiry(expiry_raw)

                is_active_raw = self._pick_value(row, "is_active", "active", "enabled")
                is_active = self._parse_bool(is_active_raw, default=True)

                minimum_raw = self._pick_value(
                    row,
                    "minimum_order_amount",
                    "minimum_order",
                    "min_order",
                    "min_order_amount",
                    "minimum",
                )
                minimum_order_amount = self._parse_decimal(minimum_raw)
                if minimum_order_amount is None:
                    minimum_order_amount = Decimal("0")

                existing = PromoCode.objects.filter(event=event, code__iexact=code).first()
                if existing:
                    existing.code = code
                    existing.discount_type = discount_type
                    existing.discount_value = discount_value
                    existing.usage_limit = usage_limit
                    existing.expiry = expiry
                    existing.is_active = is_active
                    existing.minimum_order_amount = minimum_order_amount
                    existing.save()
                    updated += 1
                else:
                    PromoCode.objects.create(
                        event=event,
                        code=code,
                        discount_type=discount_type,
                        discount_value=discount_value,
                        usage_limit=usage_limit,
                        expiry=expiry,
                        is_active=is_active,
                        minimum_order_amount=minimum_order_amount,
                    )
                    created += 1
            except Exception as exc:
                skipped += 1
                errors.append({"row": index, "error": str(exc)})

        return Response(
            {
                "total": len(rows),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )


class PromoCodeDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = PromoCodeSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrganizer]
    lookup_url_kwarg = "promo_id"
    lookup_field = "id"

    def get_queryset(self):
        event = get_object_or_404(Event, slug=self.kwargs["slug"], organizer=self.request.user)
        return event.promo_codes.all()


class PromoCodeValidateView(generics.GenericAPIView):
    serializer_class = PromoCodeValidateSerializer
    permission_classes = [permissions.AllowAny]

    @ratelimit(key="user_or_ip", rate="10/m", block=True)
    def post(self, request, *args, **kwargs):
        event = get_object_or_404(Event, slug=self.kwargs["slug"], status="published")
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        promo: PromoCode = serializer.validated_data["promo"]
        discount_amount: Decimal = serializer.validated_data["discount_amount"]
        subtotal: Decimal = serializer.validated_data["subtotal"]
        final_price = subtotal - discount_amount
        return Response(
            {
                "discount_amount": discount_amount,
                "discount_type": promo.discount_type,
                "final_price": final_price,
            },
            status=status.HTTP_200_OK,
        )
