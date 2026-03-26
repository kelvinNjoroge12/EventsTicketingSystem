from __future__ import annotations

import csv
import io

from django.core.exceptions import ValidationError
from django.db import transaction

from .models import Course, School


def _normalize_header(value) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _read_rows(uploaded_file):
    rows = []
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        content = uploaded_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            normalized = {_normalize_header(key): value for key, value in row.items() if key}
            if any(str(value).strip() for value in normalized.values() if value is not None):
                rows.append(normalized)
        return rows

    if filename.endswith(".xlsx"):
        try:
            import openpyxl
        except Exception as exc:  # pragma: no cover - import error path
            raise ValidationError("Excel upload requires openpyxl.") from exc

        workbook = openpyxl.load_workbook(uploaded_file)
        worksheet = workbook.active
        headers = [_normalize_header(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        for row in worksheet.iter_rows(min_row=2):
            data = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                data[header] = row[index].value
            if any(str(value).strip() for value in data.values() if value is not None):
                rows.append(data)
        return rows

    raise ValidationError("Upload a CSV or XLSX file.")


def _parse_sort_order(value):
    if value in (None, ""):
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ValidationError("Sort order must be a whole number.") from exc


def _string_value(row: dict, *keys) -> str:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return str(row[key]).strip()
    return ""


def import_schools_file(uploaded_file) -> dict:
    rows = _read_rows(uploaded_file)
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_number, row in enumerate(rows, start=2):
        try:
            name = _string_value(row, "name")
            if not name:
                raise ValidationError("Missing school name.")

            code = _string_value(row, "code")
            sort_order = _parse_sort_order(row.get("sort_order"))

            with transaction.atomic():
                school = School.objects.filter(name__iexact=name).first()
                if school is None:
                    School.objects.create(
                        name=name,
                        code=code,
                        is_active=True,
                        sort_order=sort_order if sort_order is not None else 0,
                    )
                    created += 1
                    continue

                update_fields = []
                if school.name != name:
                    school.name = name
                    update_fields.append("name")
                if code and school.code != code:
                    school.code = code
                    update_fields.append("code")
                if sort_order is not None and school.sort_order != sort_order:
                    school.sort_order = sort_order
                    update_fields.append("sort_order")
                if not school.is_active:
                    school.is_active = True
                    update_fields.append("is_active")
                if update_fields:
                    school.save(update_fields=update_fields)
                    updated += 1
        except ValidationError as exc:
            skipped += 1
            errors.append({"row": row_number, "error": exc.messages[0]})

    return {
        "total": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


def import_courses_file(uploaded_file) -> dict:
    rows = _read_rows(uploaded_file)
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_number, row in enumerate(rows, start=2):
        try:
            name = _string_value(row, "name")
            if not name:
                raise ValidationError("Missing course name.")

            code = _string_value(row, "code")
            school_name = _string_value(row, "school")
            sort_order = _parse_sort_order(row.get("sort_order"))

            with transaction.atomic():
                school = None
                if school_name:
                    school = School.objects.filter(name__iexact=school_name).first()
                    if school is None:
                        school = School.objects.create(name=school_name, is_active=True)
                    elif not school.is_active:
                        school.is_active = True
                        school.save(update_fields=["is_active"])

                course = Course.objects.filter(school=school, name__iexact=name).first()
                if course is None:
                    Course.objects.create(
                        school=school,
                        name=name,
                        code=code,
                        is_active=True,
                        sort_order=sort_order if sort_order is not None else 0,
                    )
                    created += 1
                    continue

                update_fields = []
                if course.name != name:
                    course.name = name
                    update_fields.append("name")
                if code and course.code != code:
                    course.code = code
                    update_fields.append("code")
                if sort_order is not None and course.sort_order != sort_order:
                    course.sort_order = sort_order
                    update_fields.append("sort_order")
                if course.school_id != getattr(school, "id", None):
                    course.school = school
                    update_fields.append("school")
                if not course.is_active:
                    course.is_active = True
                    update_fields.append("is_active")
                if update_fields:
                    course.save(update_fields=update_fields)
                    updated += 1
        except ValidationError as exc:
            skipped += 1
            errors.append({"row": row_number, "error": exc.messages[0]})

    return {
        "total": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }
